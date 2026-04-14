from pathlib import Path

from openpyxl import load_workbook

from players_contract.export.contract_salary_mapper import (
    HEBREW_HEADERS,
    ContractSalaryMapper,
)


def _sample_record() -> dict:
    return {
        "player_id": "123456",
        "team_name": "מכבי תל אביב",
        "season": "2025/26",
        "person_type": "player",
        "person_role": None,
        "employment_months": 12,
        "base_salary_monthly": 10000.0,
        "base_salary_yearly": 120000.0,
        "housing_allowance_monthly": 2000.0,
        "housing_allowance_yearly": 24000.0,
        "car_allowance_monthly": 1500.0,
        "car_allowance_yearly": 18000.0,
        "points_bonus_per_point": 500.0,
        "max_points_for_bonus": 30.0,
        "points_bonus_total": 15000.0,
        "goal_assist_penalty_bonus": 1000.0,
        "source_file": "contract.pdf",
        "page_in_document": 1,
    }


def test_writes_hebrew_headers_and_rtl(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    mapper = ContractSalaryMapper()

    result = mapper.write([_sample_record()], output)

    assert result == output
    workbook = load_workbook(result)
    worksheet = workbook.active
    assert worksheet.sheet_view.rightToLeft is True

    headers = [cell.value for cell in worksheet[1]]
    assert headers == list(HEBREW_HEADERS.values())

    row = [cell.value for cell in worksheet[2]]
    assert row[0] == "123456"
    assert row[1] == "מכבי תל אביב"
    assert row[2] == "2025/26"
    assert row[3] == "שחקן"
    assert row[4] is None
    assert row[6] == 10000.0


def test_translates_person_type_enum_values(tmp_path: Path):
    mapper = ContractSalaryMapper()
    records = [
        _sample_record() | {"person_type": "player"},
        _sample_record() | {"person_type": "coach"},
        _sample_record() | {"person_type": "other"},
    ]
    output = tmp_path / "types.xlsx"

    mapper.write(records, output)

    worksheet = load_workbook(output).active
    person_type_col = list(HEBREW_HEADERS.keys()).index("person_type") + 1
    assert worksheet.cell(row=2, column=person_type_col).value == "שחקן"
    assert worksheet.cell(row=3, column=person_type_col).value == "מאמן"
    assert worksheet.cell(row=4, column=person_type_col).value == "אחר"


def test_empty_records_writes_headers_only(tmp_path: Path):
    output = tmp_path / "empty.xlsx"
    mapper = ContractSalaryMapper()

    result = mapper.write([], output)

    workbook = load_workbook(result)
    worksheet = workbook.active
    assert worksheet.max_row == 1
    headers = [cell.value for cell in worksheet[1]]
    assert headers == list(HEBREW_HEADERS.values())


def test_forces_xlsx_suffix(tmp_path: Path):
    output_without_suffix = tmp_path / "out"
    mapper = ContractSalaryMapper()

    result = mapper.write([_sample_record()], output_without_suffix)

    assert result.suffix == ".xlsx"
    assert result.exists()


def test_handles_multiple_records(tmp_path: Path):
    output = tmp_path / "multi.xlsx"
    mapper = ContractSalaryMapper()
    records = [_sample_record(), _sample_record() | {"player_id": "999"}]

    mapper.write(records, output)

    workbook = load_workbook(output)
    worksheet = workbook.active
    assert worksheet.max_row == 3
    assert worksheet.cell(row=2, column=1).value == "123456"
    assert worksheet.cell(row=3, column=1).value == "999"


def test_header_styling(tmp_path: Path):
    output = tmp_path / "style.xlsx"
    ContractSalaryMapper().write([_sample_record()], output)

    worksheet = load_workbook(output).active
    header_cell = worksheet.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.font.color.rgb.endswith("FFFFFF")
    assert header_cell.fill.start_color.rgb.endswith("305496")
    assert header_cell.alignment.horizontal == "center"
    assert worksheet.row_dimensions[1].height == 30


def test_freeze_panes_and_autofilter(tmp_path: Path):
    output = tmp_path / "freeze.xlsx"
    ContractSalaryMapper().write([_sample_record()], output)

    worksheet = load_workbook(output).active
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions


def test_empty_records_no_autofilter(tmp_path: Path):
    output = tmp_path / "empty_filter.xlsx"
    ContractSalaryMapper().write([], output)

    worksheet = load_workbook(output).active
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref is None


def test_currency_number_format_applied(tmp_path: Path):
    output = tmp_path / "format.xlsx"
    ContractSalaryMapper().write([_sample_record()], output)

    worksheet = load_workbook(output).active
    fields = list(HEBREW_HEADERS.keys())
    salary_col = fields.index("base_salary_monthly") + 1
    months_col = fields.index("employment_months") + 1
    name_col = fields.index("team_name") + 1

    assert worksheet.cell(row=2, column=salary_col).number_format == "#,##0"
    assert worksheet.cell(row=2, column=months_col).number_format == "0"
    assert worksheet.cell(row=2, column=name_col).number_format == "General"
