from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_mapper import ExcelMapper

HEBREW_HEADERS: dict[str, str] = {
    "player_id": "מספר שחקן",
    "team_name": "שם קבוצה",
    "season": "עונה",
    "person_type": "סוג",
    "person_role": "תפקיד",
    "employment_months": "חודשי העסקה",
    "base_salary_monthly": "שכר בסיס חודשי",
    "base_salary_yearly": "שכר בסיס שנתי",
    "housing_allowance_monthly": "קצובת דיור חודשית",
    "housing_allowance_yearly": "קצובת דיור שנתית",
    "car_allowance_monthly": "קצובת רכב חודשית",
    "car_allowance_yearly": "קצובת רכב שנתית",
    "points_bonus_per_point": "בונוס לנקודה",
    "max_points_for_bonus": "נקודות מקסימום לבונוס",
    "points_bonus_total": "סך בונוס נקודות",
    "goal_assist_penalty_bonus": "בונוס שערים/בישולים/פנדלים",
    "source_file": "קובץ מקור",
    "page_in_document": "עמוד במסמך",
}

HEBREW_VALUES: dict[str, dict[str, str]] = {
    "person_type": {
        "player": "שחקן",
        "coach": "מאמן",
        "other": "אחר",
    },
}

CURRENCY_FORMAT = "#,##0"
INTEGER_FORMAT = "0"

NUMBER_FORMATS: dict[str, str] = {
    "employment_months": INTEGER_FORMAT,
    "max_points_for_bonus": INTEGER_FORMAT,
    "page_in_document": INTEGER_FORMAT,
    "base_salary_monthly": CURRENCY_FORMAT,
    "base_salary_yearly": CURRENCY_FORMAT,
    "housing_allowance_monthly": CURRENCY_FORMAT,
    "housing_allowance_yearly": CURRENCY_FORMAT,
    "car_allowance_monthly": CURRENCY_FORMAT,
    "car_allowance_yearly": CURRENCY_FORMAT,
    "points_bonus_per_point": CURRENCY_FORMAT,
    "points_bonus_total": CURRENCY_FORMAT,
    "goal_assist_penalty_bonus": CURRENCY_FORMAT,
}

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ROW_HEIGHT = 30

DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGNMENT = Alignment(horizontal="right", vertical="center")

COLUMN_WIDTH_PADDING = 3
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 28


def _translate(field: str, value):
    if value is None:
        return None
    translations = HEBREW_VALUES.get(field)
    if translations is None:
        return value
    return translations.get(value, value)


def _display_width(field: str, value) -> int:
    if value is None:
        return 0
    fmt = NUMBER_FORMATS.get(field)
    if isinstance(value, (int, float)):
        if fmt == CURRENCY_FORMAT:
            return len(f"{value:,.0f}")
        if fmt == INTEGER_FORMAT:
            return len(f"{int(value)}")
    return len(str(value))


class ContractSalaryMapper(ExcelMapper):
    def write(self, records: list[dict], output_path: Path) -> Path:
        output_path = Path(output_path).with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.sheet_view.rightToLeft = True

        fields = list(HEBREW_HEADERS.keys())
        worksheet.append([HEBREW_HEADERS[f] for f in fields])

        translated_rows = [
            [_translate(field, record.get(field)) for field in fields]
            for record in records
        ]
        for row in translated_rows:
            worksheet.append(row)

        self._style_header(worksheet, fields)
        self._style_data(worksheet, fields, len(translated_rows))
        self._set_column_widths(worksheet, fields, translated_rows)
        self._freeze_and_filter(worksheet, has_records=bool(translated_rows))

        workbook.save(output_path)
        return output_path

    def _style_header(self, worksheet, fields: list[str]) -> None:
        for col_idx in range(1, len(fields) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
        worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT

    def _style_data(self, worksheet, fields: list[str], row_count: int) -> None:
        for col_idx, field in enumerate(fields, start=1):
            number_format = NUMBER_FORMATS.get(field)
            for row_idx in range(2, 2 + row_count):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGNMENT
                if number_format is not None:
                    cell.number_format = number_format

    def _set_column_widths(self, worksheet, fields: list[str], rows: list[list]) -> None:
        for col_idx, field in enumerate(fields, start=1):
            header_width = len(HEBREW_HEADERS[field])
            value_widths = [
                _display_width(field, row[col_idx - 1])
                for row in rows
                if row[col_idx - 1] is not None
            ]
            max_value_width = max(value_widths, default=0)
            width = max(header_width, max_value_width) + COLUMN_WIDTH_PADDING
            width = max(MIN_COLUMN_WIDTH, min(width, MAX_COLUMN_WIDTH))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _freeze_and_filter(self, worksheet, has_records: bool) -> None:
        worksheet.freeze_panes = "A2"
        if has_records:
            worksheet.auto_filter.ref = worksheet.dimensions
